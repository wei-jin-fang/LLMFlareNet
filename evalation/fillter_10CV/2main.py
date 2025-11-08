import argparse
import sys

import time
import openpyxl
import torch
import torch.optim as optim
import torch.nn.functional as F
import numpy as np
import argparse
import os
import pandas as pd
from sklearn.utils import compute_class_weight

from models.LSTM import LSTMModel
from models.NN import NNModel
from models.Onefitall_16 import Onefitall_16Model
from models.Onefitall_18 import Onefitall_18Model
from models.Onefitall_26 import Onefitall_26Model
from models.Transformer import TransformerModel
from tools import BS_BSS_score, BSS_eval_np, get_batches_all

from tools import Metric, plot_losses
from tools import getClass
from tools import shuffle_data
from tools import get_batches_integer

from tools import Rectify_binary
from tools import save_torchModel
from tools import setup_seed_torch
from tools import DualOutput

'''
全局变量设置：
'''
TIME_STEPS = 40
INPUT_SIZE = 10
Class_NUM = 2
FIRST = 1

os.environ['CURL_CA_BUNDLE'] = ''
os.environ["PYTORCH_CUDA_ALLOC_CONF"] = "max_split_size_mb:64"


def load_data(train_csv_path, validate_csv_path, test_csv_path, Class_NUM):
    pd.set_option('display.max_columns', None)  # 意思是不限制显示的列数。这样设置后，无论 Pandas 数据帧有多少列
    List = []
    count_index = 0
    for path in [train_csv_path, validate_csv_path, test_csv_path]:
        count_index += 1
        csv = pd.read_csv(path)
        start = 0
        end = 0
        for Column in csv.columns.values:
            start += 1
            if Column.__eq__("TOTUSJH"):
                break

        for Column in csv.columns.values:
            end += 1
            if Column.__eq__("SHRGT45"):
                break
        List.append(csv.iloc[:, start - 1:end].values)  # train_x  test_x
        Classes = csv["CLASS"].copy()
        # 定义分类数组
        categories = ["NC", "MX"]
        # 初始化类别计数列表
        weights = [0] * len(categories)
        class_list = []

        # 遍历每个Class，计算每个类别的数量
        for Class_ in Classes:
            # 针对每一个数据
            for i, category in enumerate(categories):
                # 判断数据那个类别
                if Class_ in category:
                    weights[i] += 1
                    class_list.append(i)
                    break  # 找到匹配的类别后退出内层循环

        class_tensor = torch.tensor(class_list, dtype=torch.long)
        List.append(np.array(class_tensor))

        print(f"{path}每一类的数量是:", weights)
        tempweight = []

        for weight in weights:
            tempweight.append(weight)

        weight_list = getClass(tempweight)
        print(f"{path}get_Class函数得到的的权重：", weight_list)  # 用于测试

    return List[0], List[1], List[2], List[3], List[4], List[5]


def Preprocess(train_csv_path, validate_csv_path, test_csv_path):
    global FIRST

    train_x, train_y, validate_x, validate_y, test_x, test_y = \
        load_data(train_csv_path, validate_csv_path, test_csv_path, Class_NUM)

    train_x = train_x.reshape(-1, TIME_STEPS, INPUT_SIZE)
    train_y = Rectify_binary(train_y, Class_NUM, TIME_STEPS)
    validate_x = validate_x.reshape(-1, TIME_STEPS, INPUT_SIZE)
    test_x = test_x.reshape(-1, TIME_STEPS, INPUT_SIZE)
    validate_y = Rectify_binary(validate_y, Class_NUM, TIME_STEPS)
    test_y = Rectify_binary(test_y, Class_NUM, TIME_STEPS)

    if FIRST == 1:
        print("train_x.shape : {} ".format(train_x.shape))
        print("train_y.shape : {} ".format(train_y.shape))
        print("validate_x.shape : {} ".format(validate_x.shape))
        print("validate_y.shape : {} ".format(validate_y.shape))
        print("test_x.shape : {} ".format(test_x.shape))
        print("test_y.shape : {} ".format(test_y.shape))
        FIRST = 0

    def class_weight(alpha, beta, zero=True):
        # 权重计算-修改位置开始
        classes = np.array([0, 1])
        # 权重计算-修改位置结束
        weight = compute_class_weight(class_weight='balanced', classes=classes, y=train_y)
        if zero:
            weight[0] = weight[0] * alpha
            print("zero:", weight[0])
            return weight[0]
        else:
            weight[1] = weight[1] * beta
            print("zero:", weight[1])
            return weight[1]

    class_weight = {0.: class_weight(alpha=1, beta=1, zero=True), 1.: class_weight(alpha=1, beta=1, zero=False)}
    print(class_weight)

    return train_x, train_y, validate_x, validate_y, test_x, test_y, class_weight


def train_integer(ep, train_x, train_y, optimizer, model, batch_size):
    global steps
    train_loss = 0
    model.train()
    batch_count = 0
    for batch_idx, (data, target) in enumerate(get_batches_integer(train_x, train_y, batch_size)):
        if not isinstance(data, torch.Tensor):
            data = torch.tensor(data, dtype=torch.float32)
        # 修改 target 处理：无论是否为 Tensor，都强制转换为 float32
        target = torch.tensor(target, dtype=torch.float32) if not isinstance(target, torch.Tensor) else target.to(
            dtype=torch.float32)

        if args.cuda:
            data, target = data.cuda(), target.cuda()
            class_weights_cuda = torch.tensor([class_weight[0.], class_weight[1.]], dtype=torch.float32).cuda()
        else:
            class_weights_cuda = torch.tensor([class_weight[0.], class_weight[1.]], dtype=torch.float32)

        optimizer.zero_grad()
        output = model(data)  # [batch_size, 1]，概率值
        target = target.view(-1, 1)  # 确保 target 形状为 [batch_size, 1]
        # 根据 target 生成每个样本的权重
        weights = torch.zeros_like(target, dtype=torch.float32, device=target.device)
        weights[target == 0] = class_weights_cuda[0]  # 类别 0 的权重
        weights[target == 1] = class_weights_cuda[1]  # 类别 1 的权重
        loss = F.binary_cross_entropy(output, target, weight=weights)  # 使用 BCE Loss
        loss.backward()
        optimizer.step()
        train_loss += loss.item()
        batch_count += 1

    message = ('Train Epoch: {} \t average Loss: {:.6f}'.format(ep, train_loss / batch_count))
    print(message)
    return train_loss / batch_count


def train_all(ep, train_x, train_y, optimizer, model, batch_size):
    global steps
    train_loss = 0
    model.train()
    batch_count = 0
    for batch_idx, (data, target) in enumerate(get_batches_all(train_x, train_y, batch_size)):
        if not isinstance(data, torch.Tensor):
            data = torch.tensor(data, dtype=torch.float32)
        # 修改 target 处理：无论是否为 Tensor，都强制转换为 float32
        target = torch.tensor(target, dtype=torch.float32) if not isinstance(target, torch.Tensor) else target.to(
            dtype=torch.float32)

        if args.cuda:
            data, target = data.cuda(), target.cuda()
            class_weights_cuda = torch.tensor([class_weight[0.], class_weight[1.]], dtype=torch.float32).cuda()
        else:
            class_weights_cuda = torch.tensor([class_weight[0.], class_weight[1.]], dtype=torch.float32)

        optimizer.zero_grad()
        output = model(data)  # [batch_size, 1]，概率值
        target = target.view(-1, 1)  # 确保 target 形状为 [batch_size, 1]
        # 根据 target 生成每个样本的权重
        weights = torch.zeros_like(target, dtype=torch.float32, device=target.device)
        weights[target == 0] = class_weights_cuda[0]  # 类别 0 的权重
        weights[target == 1] = class_weights_cuda[1]  # 类别 1 的权重
        loss = F.binary_cross_entropy(output, target, weight=weights)  # 使用 BCE Loss
        loss.backward()
        optimizer.step()
        train_loss += loss.item()
        batch_count += 1

    message = ('Train Epoch: {} \t average Loss: {:.6f}'.format(ep, train_loss / batch_count))
    print(message)
    return train_loss / batch_count


def evaluate_integer(data_x, data_y, model, batch_size):
    model.eval()
    test_loss = 0
    correct = 0
    all_targets = []
    all_predictions = []
    batch_count = 0
    all_predictions_y_true = []
    all_predictions_y_prob = []

    with torch.no_grad():
        for data, target in get_batches_integer(data_x, data_y, batch_size):
            if not isinstance(data, torch.Tensor):
                data = torch.tensor(data, dtype=torch.float32)
            # 修改 target 处理：无论是否为 Tensor，都强制转换为 float32
            target = torch.tensor(target, dtype=torch.float32) if not isinstance(target, torch.Tensor) else target.to(
                dtype=torch.float32)

            if args.cuda:
                data, target = data.cuda(), target.cuda()
                class_weights_cuda = torch.tensor([class_weight[0.], class_weight[1.]], dtype=torch.float32).cuda()
            else:
                class_weights_cuda = torch.tensor([class_weight[0.], class_weight[1.]], dtype=torch.float32)

            output = model(data)  # [batch_size, 1]，概率值
            target = target.view(-1, 1)
            # 根据 target 生成每个样本的权重
            weights = torch.zeros_like(target, dtype=torch.float32, device=target.device)
            weights[target == 0] = class_weights_cuda[0]  # 类别 0 的权重
            weights[target == 1] = class_weights_cuda[1]  # 类别 1 的权重
            test_loss += F.binary_cross_entropy(output, target, weight=weights)  # 使用 BCE Loss.item()

            pred = (output > 0.5).float()  # 阈值 0.5 这个地方取值就是0或者1

            all_predictions.extend(pred.cpu().numpy().flatten())  # 对应之前的：这一批次预测laebl加进去数组
            correct += pred.eq(target.data.view_as(pred)).cpu().sum()
            all_targets.extend(target.cpu().numpy().flatten())  # 对应之前的：这一批次实际label加进去数组

            # 添加内容方便计算 BSS BS
            all_predictions_y_true.extend(target.cpu().numpy().flatten())  # 对应之前的：拿到实际的label
            pos_prob = output  # 正类概率 [batch_size, 1]
            neg_prob = 1.0 - pos_prob  # 负类概率 [batch_size, 1]
            probabilities = torch.cat((neg_prob, pos_prob), dim=1)  # [batch_size, 2]，[负类概率, 正类概率]
            all_predictions_y_prob.extend(probabilities.cpu().numpy().tolist())  # 存储 [负类概率, 正类概率] 列表 #对应之前的： 拿到概率数值加入数组

            batch_count += 1

        metric = Metric(all_targets, all_predictions)
        print(metric.Matrix())
        TSS = metric.TSS()[0]
        print(f"TSS: {TSS}")
        test_loss /= batch_count
        accuracy = correct / len(data_x)
        message = (
            '\nTesting: Average loss: {:.4f}, Accuracy: {}/{} ({:.0f}%)\n'.format(
                test_loss, correct, len(data_x), 100. * accuracy))
        print("本轮评估完成", message)
        return {'loss': test_loss, 'accuracy': accuracy, 'tss': TSS, "metric": metric}, \
               all_predictions_y_true, all_predictions_y_prob


def evalual_all(data_x, data_y, model, batch_size):
    model.eval()
    test_loss = 0
    correct = 0
    all_targets = []
    all_predictions = []
    batch_count = 0
    all_predictions_y_true = []
    all_predictions_y_prob = []

    with torch.no_grad():
        for data, target in get_batches_all(data_x, data_y, batch_size):
            if not isinstance(data, torch.Tensor):
                data = torch.tensor(data, dtype=torch.float32)
            # 修改 target 处理：无论是否为 Tensor，都强制转换为 float32
            target = torch.tensor(target, dtype=torch.float32) if not isinstance(target, torch.Tensor) else target.to(
                dtype=torch.float32)

            if args.cuda:
                data, target = data.cuda(), target.cuda()
                class_weights_cuda = torch.tensor([class_weight[0.], class_weight[1.]], dtype=torch.float32).cuda()
            else:
                class_weights_cuda = torch.tensor([class_weight[0.], class_weight[1.]], dtype=torch.float32)

            output = model(data)  # [batch_size, 1]，概率值
            target = target.view(-1, 1)
            # 根据 target 生成每个样本的权重
            weights = torch.zeros_like(target, dtype=torch.float32, device=target.device)
            weights[target == 0] = class_weights_cuda[0]  # 类别 0 的权重
            weights[target == 1] = class_weights_cuda[1]  # 类别 1 的权重
            test_loss += F.binary_cross_entropy(output, target, weight=weights)

            pred = (output > 0.5).float()  # 阈值 0.5 这个地方取值就是0或者1

            all_predictions.extend(pred.cpu().numpy().flatten())  # 对应之前的：这一批次预测laebl加进去数组
            correct += pred.eq(target.data.view_as(pred)).cpu().sum()
            all_targets.extend(target.cpu().numpy().flatten())  # 对应之前的：这一批次实际label加进去数组

            # 添加内容方便计算 BSS BS
            all_predictions_y_true.extend(target.cpu().numpy().flatten())  # 对应之前的：拿到实际的label
            pos_prob = output  # 正类概率 [batch_size, 1]
            neg_prob = 1.0 - pos_prob  # 负类概率 [batch_size, 1]
            probabilities = torch.cat((neg_prob, pos_prob), dim=1)  # [batch_size, 2]，[负类概率, 正类概率]
            all_predictions_y_prob.extend(probabilities.cpu().numpy().tolist())  # 存储 [负类概率, 正类概率] 列表 #对应之前的： 拿到概率数值加入数组

            batch_count += 1

        metric = Metric(all_targets, all_predictions)
        print(metric.Matrix())
        TSS = metric.TSS()[0]
        print(f"TSS: {TSS}")
        test_loss /= batch_count
        accuracy = correct / len(data_x)
        message = (
            '\nTesting: Average loss: {:.4f}, Accuracy: {}/{} ({:.0f}%)\n'.format(
                test_loss, correct, len(data_x), 100. * accuracy))
        print("本轮评估完成", message)
        return {'loss': test_loss, 'accuracy': accuracy, 'tss': TSS, "metric": metric}, \
               all_predictions_y_true, all_predictions_y_prob




def save_args_to_csv(args, model_base):
    """将args参数保存到CSV文件中，方便调参记录"""
    args_dict = vars(args)  # 将args转换为字典

    # 添加时间戳
    args_dict['timestamp'] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    args_dict['model_output_path'] = model_base

    # 创建DataFrame
    df = pd.DataFrame([args_dict])

    # 保存到model_base目录下
    csv_path = f"{model_base}/args_config.csv"
    df.to_csv(csv_path, index=False)
    print(f"参数配置已保存到: {csv_path}")

    # 同时保存到项目根目录的汇总文件
    summary_csv = "args_history.csv"
    if os.path.exists(summary_csv):
        # 如果文件存在，追加数据
        df.to_csv(summary_csv, mode='a', header=False, index=False)
    else:
        # 如果文件不存在，创建新文件
        df.to_csv(summary_csv, index=False)
    print(f"参数配置已追加到汇总文件: {summary_csv}")

    return csv_path


device = torch.device("cuda" if torch.cuda.is_available() else "cpu")


def get_model(model_name):
    # 在主程序中定义模型映射
    model_dict = {
        # "LLMFlareNet_1": LLMFlareNet_1Model,
        # "LLMFlareNet_2": LLMFlareNet_2Model,
        # "LLMFlareNet_6": LLMFlareNet_6Model,
        # "LLMFlareNet_5": LLMFlareNet_5Model,

        "Onefitall_16": Onefitall_16Model,
        "Onefitall_18": Onefitall_18Model,
        "Onefitall_26": Onefitall_26Model,
        "Transformer": TransformerModel,
        "LSTM": LSTMModel,
        "NN": NNModel,
    }

    # 实例化模型
    model_class = model_dict.get(model_name)
    if model_class is None:
        raise ValueError(f"Unknown model type: {args.model_type}")
    return model_class


def read_parameters():
    # 初始化参数收集器
    parser = argparse.ArgumentParser(description='Time-LLM')
    # 数据集部分参数
    parser.add_argument('--num_dataset', type=int, default=9)  # 循环处理0-9个数据集
    parser.add_argument('--input_size', type=int, default=10)  # 循环处理0-9个数据集
    parser.add_argument('--time_step', type=int, default=40)  # 循环处理0-9个数据集
    parser.add_argument('--batch_size', type=int, default=16, help='batch size of train input data')
    parser.add_argument('--device', type=str, default="cuda")  # 循环处理0-9个数据集
    parser.add_argument('--epochs', type=int, default=50)
    # 公共训练参数
    parser.add_argument('--cuda', action='store_false')
    parser.add_argument('--lr', type=float, default=0.001)
    parser.add_argument('--optim', type=str, default='Adam')
    parser.add_argument('--seed', type=int, default=2021, help='random seed')

    parser.add_argument('--datasetname', type=str, default="data_fillter_single", help='data_fillter_single,data_fillter_mutiple')
    parser.add_argument('--model_type', type=str, default='NN',help='Onefitall,LLMFlareNet')
    args = parser.parse_args()
    return args

if __name__ == "__main__":

    args = read_parameters()
    batch_size=16
    start_time = time.time()
    timelabel = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime(time.time()))
    # 定义训练输出目录
    model_base = f"./model_output/{timelabel}"
    os.makedirs(f"{model_base}")
    os.makedirs(f"{model_base}/plot")
    results_filepath = f"{model_base}/important.txt"
    results_logfilepath = f"{model_base}/log.txt"
    # 保证输出到txt和控制台
    sys.stdout = DualOutput(results_logfilepath)

    # 定义十个数据集合存储内容
    all_matrix = np.array([[0, 0], [0, 0]])
    data_Recall, data_Precision, data_Accuracy, data_TSS, data_BSS, data_HSS, data_FAR = [], [], [], [], [], [], []

    # 打开文件准备写入
    with open(results_filepath, 'w') as results_file:
        for count in range(args.num_dataset + 1):  # 循环处理0-9个数据集
            setup_seed_torch(args.seed)

            train_csv_path = f"./{args.datasetname}/{count}Test.csv"
            validate_csv_path = train_csv_path
            test_csv_path = train_csv_path

            train_x, train_y, validate_x, validate_y, test_x, test_y, class_weight = \
                Preprocess(train_csv_path, validate_csv_path, test_csv_path)

            print(f"====================数据集{count}测试集轮评估数据=============================================")
            # 加载最佳模型
            model_test = torch.load(f"../../weight/NN/model_{count}.pt")
            # 测试集评估
            test_metrics, all_predictions_y_true, all_predictions_y_prob = \
                evalual_all(test_x, test_y, model_test, batch_size)
            del model_test
            # 清理 GPU 内存
            # torch.cuda.empty_cache()

            # 计算测试集矩阵
            testMetrics = test_metrics["metric"]
            print(testMetrics.Matrix())
            all_matrix += testMetrics.Matrix()

            data_Recall.append(testMetrics.Recall())
            print("Recall", testMetrics.Recall())

            data_Precision.append(testMetrics.Precision())
            print("Precision", testMetrics.Precision())

            data_Accuracy.append(testMetrics.Accuracy())
            print("Accuracy", testMetrics.Accuracy())

            data_TSS.append(testMetrics.TSS())
            print("TSS", testMetrics.TSS())

            data_HSS.append(testMetrics.HSS())
            print("HSS", testMetrics.HSS())

            data_FAR.append(testMetrics.FAR())
            print("FAR", testMetrics.FAR())

            # 开始求BSS
            y_true = all_predictions_y_true
            y_prob = np.array([row[1] for row in all_predictions_y_prob])

            BS, BSS = BS_BSS_score(y_true, y_prob)
            data_BSS.append([BS, BSS])
            print("BS, BSS", [BS, BSS])

            print(f"数据集 {count} 测试集TSS:", test_metrics['tss'])
            # 写入结果到文件
            results_file.write(f"数据集 {count} 测试集TSS: {test_metrics['tss']}\n")
            results_file.write("=================================================================\n")
        print("#接下来计算所有测试集指标均值和方法")
        print(all_matrix)
        print(data_BSS)
        # 转换数据为numpy数组以便计算
        data_Recall = np.array(data_Recall)
        data_Precision = np.array(data_Precision)
        data_Accuracy = np.array(data_Accuracy)
        data_TSS = np.array(data_TSS)
        data_HSS = np.array(data_HSS)
        data_FAR = np.array(data_FAR)
        data_BSS = np.array(data_BSS)
        # 计算均值和标准差
        results = {
            "Metric": ["Recall", "Precision", "Accuracy", "TSS", "HSS", "FAR", "BSS"],
            "Mean": [data_Recall.mean(axis=0), data_Precision.mean(axis=0), data_Accuracy.mean(axis=0),
                     data_TSS.mean(axis=0), data_HSS.mean(axis=0), data_FAR.mean(axis=0), data_BSS.mean(axis=0)],
            "Std": [data_Recall.std(axis=0), data_Precision.std(axis=0), data_Accuracy.std(axis=0),
                    data_TSS.std(axis=0), data_HSS.std(axis=0), data_FAR.std(axis=0), data_BSS.std(axis=0)]
        }
        print("++++++++++++++++")
        print(data_TSS.mean(axis=0))
        print("++++++++++++++++")
        # 将结果写入Excel
        df = pd.DataFrame(results)
        excel_filename = f'{model_base}/results.xlsx'
        df.to_excel(excel_filename, index=False)
        print(f"结果已写入 {excel_filename}")

        # 记录当前TSS
        # 设置汇总result.xlsx
        filename = 'result.xlsx'
        new_row = [f'{model_base}', data_TSS.mean(axis=0)[1]]

        # 判断文件是否存在
        if os.path.exists(filename):
            # 文件存在，加载文件
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
        else:
            # 文件不存在，创建一个新工作簿
            wb = openpyxl.Workbook()
            ws = wb.active

        # 获取当前行数
        current_row = ws.max_row + 1

        # 将数据写入新行
        for col_num, value in enumerate(new_row, start=1):
            ws.cell(row=current_row, column=col_num, value=value)

        # 保存文件
        wb.save(filename)

    # 删除模型
    for filename in os.listdir(model_base):
        # 检查文件是否以 .pt 结尾
        if filename.endswith('.pt'):
            # 构造文件的完整路径
            file_path = os.path.join(model_base, filename)
            try:
                # 删除文件
                os.remove(file_path)
                pass
            except Exception as e:
                print(f"删除文件 {file_path} 时出错: {e}")

    end_time = time.time()
    elapsed_time_minutes = (end_time - start_time) / 60
    print(f"程序运行时间: {elapsed_time_minutes:.2f} 分钟")

